#!/usr/bin/env python3
"""
Content Analysis Report Generator — Non-destructive review.

Does NOT change content. Reads all sections and produces a
spreadsheet with detailed findings + observations per section.

Uses evaluation board + gate filtering to score each section,
then logs everything to an xlsx file.

Usage:
    export ANTHROPIC_API_KEY=sk-ant-...
    python3 report_generator.py --input book.docx
    python3 agent_loop.py --mode review (preferred)

Legacy tool. Use agent_loop.py instead.
Framework: Lyra Labs, 2026
"""

import json
import os
import sys
import re
import time
import argparse
import subprocess
from pathlib import Path
from datetime import datetime, timezone

sys.path.insert(0, str(Path(__file__).parent))

from reader_board_v2 import score_both_boards, lyra_gate, call_api, BATCHED_SYSTEM
from reader_board import INDUSTRY_TABLE, READER_TABLE
from edit_judge import quick_contamination_check, quick_voice_check
from voice_fingerprint import VOICE_SAMPLE, BANNED_PATTERNS, VOICE_MARKERS
from table_read_voice import check_table_read_match

SCRIPT_DIR = Path(__file__).parent


# ═══════════════════════════════════════════════════════════
# CHAPTER EXTRACTION
# ═══════════════════════════════════════════════════════════

def extract_chapters_from_docx(docx_path):
    """Extract individual chapters from the manuscript docx."""
    result = subprocess.run(
        ["pandoc", str(docx_path), "-t", "plain", "--wrap=none"],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        print(f"ERROR: pandoc failed: {result.stderr[:200]}")
        return {}

    text = result.stdout
    lines = text.split("\n")

    # Find chapter boundaries
    chapter_pattern = re.compile(
        r'^(Prologue|Chapter\s+\d+|MOVEMENT\s+\w+\s+SUMMARY|Movement\s+\w+\s+Summary)',
        re.IGNORECASE
    )

    chapters = {}
    current_title = None
    current_lines = []

    for line in lines:
        match = chapter_pattern.match(line.strip())
        if match:
            # Save previous chapter
            if current_title and current_lines:
                text = "\n".join(current_lines).strip()
                if len(text.split()) > 50:  # Skip very short sections
                    chapters[current_title] = text
            current_title = line.strip()
            current_lines = []
        elif current_title:
            current_lines.append(line)

    # Save last chapter
    if current_title and current_lines:
        text = "\n".join(current_lines).strip()
        if len(text.split()) > 50:
            chapters[current_title] = text

    return chapters


# ═══════════════════════════════════════════════════════════
# THE FINAL EDITOR REVIEW
# ═══════════════════════════════════════════════════════════

REVIEW_SYSTEM = """You are a world-class book editor doing a FINAL review of the manuscript. You do NOT change any text. You identify issues and suggest fixes.

This is a REDUCTION review. The author wants to cut the book down — fewer chapters, less repetition, tighter field stories. Be honest about what isn't working.

For each chapter, you evaluate:

1. VOICE INTEGRITY: Does this sound like the author? Flag any sentences that sound like AI, an editor, or generic self-help.
2. CHAPTER ARCHITECTURE: Does it follow Field Insight → Field Memory → Core Truth → Framework → Standard Practice → Inquiry? Which beats are missing or weak?
3. MEME QUOTES: Are there 5+ standalone quotable lines (under 20 words, work without context)?
4. PACING: Are there pacing deaths? Long stretches without short-sentence relief?
5. CONTAMINATION: Any banned phrases, outline residue, or AI vocabulary?
6. 60 MINUTES TEST: Could a reporter film someone doing the Standard Practice?
7. REPEATABILITY: Would a reader get the same insight re-reading this in 2 years?
8. AUDIOBOOK READINESS: Can this be read aloud naturally? Any sentences that stumble?
9. FIELD STORY STRENGTH: Does every Field Memory have: (a) a named person or specific place, (b) sensory detail you can see/hear/feel, (c) body truth — a physical sensation that proves the moment was real? Flag stories that read like summaries instead of lived moments.
10. CORE THEME: What is the ONE core theme/insight of this chapter? State it in under 10 words. This will be used to detect repetition across chapters.
11. CUT ASSESSMENT: Could this chapter be:
    - ESSENTIAL (book doesn't work without it)
    - MERGEABLE (good content but overlaps with another chapter — say which one)
    - CUTTABLE (the insight exists better elsewhere in the book)
    Rate honestly. A 40-chapter book that could be 28 chapters is better for everyone.

For each issue found, provide:
- LOCATION: Quote the specific problematic text (5-10 words)
- ISSUE: What's wrong (one sentence)
- SEVERITY: HIGH / MEDIUM / LOW
- CATEGORY: voice|structure|pacing|contamination|audiobook|quotability|repeatability|field_story|repetition|cut_candidate
- SUGGESTED FIX: What the author should consider changing (specific, actionable)

Respond with ONLY JSON:
{
  "chapter_grade": "A|B|C|D|F",
  "voice_score": <1-10>,
  "structure_score": <1-10>,
  "quotability_score": <1-10>,
  "pacing_score": <1-10>,
  "audiobook_score": <1-10>,
  "field_story_score": <1-10>,
  "overall_score": <1-10>,
  "meme_quotes_found": <count>,
  "best_line": "<the single best line in the chapter>",
  "weakest_line": "<the single weakest line>",
  "core_theme": "<the ONE core theme in under 10 words>",
  "cut_recommendation": "ESSENTIAL|MERGEABLE|CUTTABLE",
  "merge_candidate": "<chapter title this could merge with, or null>",
  "cut_reasoning": "<1-2 sentences explaining the cut/merge/keep recommendation>",
  "field_stories": [
    {
      "description": "<brief description of the story>",
      "has_named_person": true|false,
      "has_sensory_detail": true|false,
      "has_body_truth": true|false,
      "landing": "LANDS|WEAK|SUMMARY",
      "note": "<what's missing or what works>"
    }
  ],
  "repeated_themes": ["<theme or phrase that appears in other chapters>"],
  "issues": [
    {
      "location": "<quoted text>",
      "issue": "<what's wrong>",
      "severity": "HIGH|MEDIUM|LOW",
      "category": "voice|structure|pacing|contamination|audiobook|quotability|repeatability|field_story|repetition|cut_candidate",
      "suggested_fix": "<specific suggestion>"
    }
  ],
  "meme_quotes": ["<quote1>", "<quote2>", ...],
  "summary": "<2-3 sentence editorial summary>"
}"""


def review_chapter(chapter_title, chapter_text, api_key, model="claude-sonnet-4-20250514"):
    """Run the full editorial review on a single chapter."""

    # Local checks first (free, instant)
    contamination = quick_contamination_check(chapter_text)
    voice = quick_voice_check(chapter_text)
    tr = check_table_read_match(chapter_text)
    gate_pass, gate_report = lyra_gate(chapter_text)
    word_count = len(chapter_text.split())

    # API review (Sonnet for quality)
    review_prompt = f"""Review this chapter from the manuscript.

CHAPTER: {chapter_title}
WORD COUNT: {word_count}

{VOICE_SAMPLE}

BANNED PATTERNS: {json.dumps(BANNED_PATTERNS[:20])}

TABLE READ VOICE CHECK:
- Spoken phrases found: {tr['spoken_phrases_found']}
- Contamination words: {tr['contamination_found'][:5] if tr['contamination_found'] else 'none'}
- Voice pair violations: {tr['voice_pair_violations'][:3] if tr['voice_pair_violations'] else 'none'}

CHAPTER TEXT:
---
{chapter_text}
---

Review every paragraph. Find EVERY issue. Be thorough and specific.
Quote the exact problematic text for each issue found."""

    try:
        text = call_api(review_prompt, REVIEW_SYSTEM, api_key, model, max_tokens=4096)
        # Parse JSON
        if text.startswith("```"):
            text = text.split("```")[1]
            if text.startswith("json"):
                text = text[4:]
            text = text.strip()
        if not text.startswith("{"):
            start = text.find("{")
            end = text.rfind("}")
            if start != -1 and end != -1:
                text = text[start:end + 1]
        review = json.loads(text)
    except Exception as e:
        print(f"    Review failed: {e}")
        review = {
            "chapter_grade": "?",
            "voice_score": 0, "structure_score": 0, "quotability_score": 0,
            "pacing_score": 0, "audiobook_score": 0, "overall_score": 0,
            "meme_quotes_found": 0, "best_line": "", "weakest_line": "",
            "issues": [], "meme_quotes": [], "summary": f"Review failed: {e}"
        }

    # Combine local + API results
    review["word_count"] = word_count
    review["contamination"] = contamination
    review["voice_check"] = voice
    review["table_read"] = {
        "score": tr["table_read_score"],
        "spoken_phrases": tr.get("spoken_phrases_found", 0),
        "contamination_words": tr.get("contamination_found", [])[:5],
        "voice_violations": tr.get("voice_pair_violations", [])[:3],
    }
    review["lyra_gate"] = gate_report.get("gate", "?")

    return review


# ═══════════════════════════════════════════════════════════
# CROSS-CHAPTER REPETITION ANALYSIS
# ═══════════════════════════════════════════════════════════

REPETITION_SYSTEM = """You are analyzing theme repetition across chapters of the manuscript. You will receive a list of chapters with their core themes.

Your job: identify which chapters share the SAME core insight, even if worded differently. Group overlapping chapters together and recommend which ones to keep, merge, or cut.

Be aggressive. A great book says each thing ONCE, in the best possible chapter. If two chapters both teach "constraints breed creativity," one of them needs to go.

Respond with ONLY JSON:
{
  "theme_clusters": [
    {
      "theme": "<shared theme in plain language>",
      "chapters": ["<chapter title 1>", "<chapter title 2>"],
      "recommendation": "<which to keep, which to merge/cut, and why>",
      "severity": "HIGH|MEDIUM|LOW"
    }
  ],
  "unique_chapters": ["<chapters that say something no other chapter says>"],
  "total_reducible": <number of chapters that could be cut or merged>,
  "recommended_final_count": <suggested total chapter count>,
  "summary": "<3-4 sentence summary of the repetition landscape>"
}"""


def analyze_cross_chapter_repetition(reviews, api_key, model="claude-sonnet-4-20250514"):
    """Compare all chapters for theme overlap and repetition."""
    chapter_summaries = []
    for title, review in reviews.items():
        core_theme = review.get("core_theme", "unknown")
        repeated = review.get("repeated_themes", [])
        cut_rec = review.get("cut_recommendation", "?")
        word_count = review.get("word_count", 0)
        grade = review.get("chapter_grade", "?")
        chapter_summaries.append(
            f"- {title} ({word_count} words, grade {grade}): "
            f"Core theme: \"{core_theme}\" | Cut rec: {cut_rec} | "
            f"Repeated themes: {', '.join(repeated) if repeated else 'none noted'}"
        )

    prompt = f"""Here are {len(reviews)} chapters from the manuscript with their core themes.
Find theme clusters — chapters that teach the same insight.

{chr(10).join(chapter_summaries)}

Be specific about which chapters overlap and what to do about it."""

    try:
        text = call_api(prompt, REPETITION_SYSTEM, api_key, model, max_tokens=4096)
        if not text.startswith("{"):
            start = text.find("{")
            end = text.rfind("}")
            if start != -1 and end != -1:
                text = text[start:end + 1]
        return json.loads(text)
    except Exception as e:
        print(f"    Cross-chapter analysis failed: {e}")
        return {
            "theme_clusters": [],
            "unique_chapters": [],
            "total_reducible": 0,
            "recommended_final_count": len(reviews),
            "summary": f"Analysis failed: {e}"
        }


# ═══════════════════════════════════════════════════════════
# SPREADSHEET OUTPUT
# ═══════════════════════════════════════════════════════════

def build_review_spreadsheet(reviews, output_path, repetition_analysis=None):
    """Build the master editorial review spreadsheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Colors
    header_fill = PatternFill("solid", fgColor="8B7EC8")  # Lyra lavender
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    grade_a = PatternFill("solid", fgColor="C6EFCE")
    grade_b = PatternFill("solid", fgColor="E2EFDA")
    grade_c = PatternFill("solid", fgColor="FFF2CC")
    grade_d = PatternFill("solid", fgColor="FCE4D6")
    grade_f = PatternFill("solid", fgColor="F4CCCC")
    essential_fill = PatternFill("solid", fgColor="C6EFCE")
    mergeable_fill = PatternFill("solid", fgColor="FFF2CC")
    cuttable_fill = PatternFill("solid", fgColor="F4CCCC")
    thin_border = Border(
        left=Side(style="thin", color="E0DDF0"),
        right=Side(style="thin", color="E0DDF0"),
        top=Side(style="thin", color="E0DDF0"),
        bottom=Side(style="thin", color="E0DDF0"),
    )

    def style_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

    # ─── Sheet 1: Chapter Scorecard ─────────────────────
    ws = wb.active
    ws.title = "Chapter Scorecard"

    headers = [
        "Chapter", "Words", "Grade", "Cut Rec", "Core Theme", "Voice",
        "Structure", "Field Story", "Quotability", "Pacing", "Audiobook",
        "Overall", "Meme Quotes", "Contamination", "Lyra Gate",
        "Best Line", "Weakest Line", "Cut Reasoning", "Summary"
    ]
    style_header(ws, headers)

    for row_idx, (title, review) in enumerate(reviews.items(), 2):
        grade = review.get("chapter_grade", "?")
        cut_rec = review.get("cut_recommendation", "?")
        grade_fill = {"A": grade_a, "B": grade_b, "C": grade_c, "D": grade_d, "F": grade_f}.get(grade, None)
        cut_fill = {"ESSENTIAL": essential_fill, "MERGEABLE": mergeable_fill, "CUTTABLE": cuttable_fill}.get(cut_rec, None)

        data = [
            title,
            review.get("word_count", 0),
            grade,
            cut_rec,
            review.get("core_theme", "")[:60],
            review.get("voice_score", 0),
            review.get("structure_score", 0),
            review.get("field_story_score", 0),
            review.get("quotability_score", 0),
            review.get("pacing_score", 0),
            review.get("audiobook_score", 0),
            review.get("overall_score", 0),
            review.get("meme_quotes_found", 0),
            len(review.get("contamination", [])),
            review.get("lyra_gate", "?"),
            review.get("best_line", "")[:100],
            review.get("weakest_line", "")[:100],
            review.get("cut_reasoning", "")[:150],
            review.get("summary", "")[:200],
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 3 and grade_fill:
                cell.fill = grade_fill
                cell.alignment = Alignment(horizontal="center")
            if col == 4 and cut_fill:
                cell.fill = cut_fill
                cell.alignment = Alignment(horizontal="center")

    # Column widths
    widths = [35, 8, 7, 11, 30, 7, 9, 10, 10, 7, 9, 8, 10, 12, 10, 40, 40, 45, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ─── Sheet 2: All Issues ─────────────────────────────
    ws2 = wb.create_sheet("All Issues")

    issue_headers = ["Chapter", "Severity", "Category", "Location", "Issue", "Suggested Fix"]
    for col, h in enumerate(issue_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    severity_fills = {
        "HIGH": PatternFill("solid", fgColor="F4CCCC"),
        "MEDIUM": PatternFill("solid", fgColor="FFF2CC"),
        "LOW": PatternFill("solid", fgColor="E2EFDA"),
    }

    issue_row = 2
    for title, review in reviews.items():
        for issue in review.get("issues", []):
            data = [
                title,
                issue.get("severity", "?"),
                issue.get("category", "?"),
                issue.get("location", "")[:80],
                issue.get("issue", ""),
                issue.get("suggested_fix", ""),
            ]
            for col, val in enumerate(data, 1):
                cell = ws2.cell(row=issue_row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if col == 2:
                    sev = issue.get("severity", "")
                    if sev in severity_fills:
                        cell.fill = severity_fills[sev]
            issue_row += 1

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 35
    ws2.column_dimensions["E"].width = 45
    ws2.column_dimensions["F"].width = 50

    # ─── Sheet 3: Meme Quotes Bank ──────────────────────
    ws3 = wb.create_sheet("Meme Quotes")

    mq_headers = ["Chapter", "Quote", "Word Count"]
    for col, h in enumerate(mq_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    mq_row = 2
    for title, review in reviews.items():
        for q in review.get("meme_quotes", []):
            ws3.cell(row=mq_row, column=1, value=title).border = thin_border
            ws3.cell(row=mq_row, column=2, value=q).border = thin_border
            ws3.cell(row=mq_row, column=3, value=len(q.split())).border = thin_border
            mq_row += 1

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 60
    ws3.column_dimensions["C"].width = 12

    # ─── Sheet 4: Voice Contamination Detail ────────────
    ws4 = wb.create_sheet("Voice Detail")

    vd_headers = ["Chapter", "Contamination Phrases", "Table Read Score",
                   "Spoken Phrases Found", "Voice Pair Violations", "Body Truth", "Short Punches"]
    for col, h in enumerate(vd_headers, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    vd_row = 2
    for title, review in reviews.items():
        voice = review.get("voice_check", {})
        tr = review.get("table_read", {})
        contam = review.get("contamination", [])
        ws4.cell(row=vd_row, column=1, value=title).border = thin_border
        ws4.cell(row=vd_row, column=2, value=", ".join(contam) if contam else "clean").border = thin_border
        ws4.cell(row=vd_row, column=3, value=tr.get("score", 0)).border = thin_border
        sp = tr.get("spoken_phrases", 0)
        ws4.cell(row=vd_row, column=4, value=len(sp) if isinstance(sp, list) else sp).border = thin_border
        vv = tr.get("voice_violations", [])
        ws4.cell(row=vd_row, column=5, value=", ".join(str(v) for v in vv) if vv else "none").border = thin_border
        bt = voice.get("body_truth", 0)
        ws4.cell(row=vd_row, column=6, value=len(bt) if isinstance(bt, list) else bt).border = thin_border
        sp2 = voice.get("short_punches", 0)
        ws4.cell(row=vd_row, column=7, value=len(sp2) if isinstance(sp2, list) else sp2).border = thin_border
        vd_row += 1

    for col_letter in ["A", "B", "C", "D", "E", "F", "G"]:
        ws4.column_dimensions[col_letter].width = 20
    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 40

    # ─── Sheet 5: Field Stories ────────────────────────
    ws5 = wb.create_sheet("Field Stories")

    fs_headers = ["Chapter", "Story", "Named Person", "Sensory Detail",
                   "Body Truth", "Landing", "Note"]
    style_header(ws5, fs_headers)

    lands_fill = PatternFill("solid", fgColor="C6EFCE")
    weak_fill = PatternFill("solid", fgColor="FFF2CC")
    summary_fill = PatternFill("solid", fgColor="F4CCCC")
    landing_fills = {"LANDS": lands_fill, "WEAK": weak_fill, "SUMMARY": summary_fill}

    fs_row = 2
    for title, review in reviews.items():
        for story in review.get("field_stories", []):
            data = [
                title,
                story.get("description", "")[:80],
                "Yes" if story.get("has_named_person") else "No",
                "Yes" if story.get("has_sensory_detail") else "No",
                "Yes" if story.get("has_body_truth") else "No",
                story.get("landing", "?"),
                story.get("note", "")[:100],
            ]
            for col, val in enumerate(data, 1):
                cell = ws5.cell(row=fs_row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if col == 6:
                    landing = story.get("landing", "")
                    if landing in landing_fills:
                        cell.fill = landing_fills[landing]
            fs_row += 1

    ws5.column_dimensions["A"].width = 30
    ws5.column_dimensions["B"].width = 40
    for c in ["C", "D", "E"]:
        ws5.column_dimensions[c].width = 14
    ws5.column_dimensions["F"].width = 10
    ws5.column_dimensions["G"].width = 45

    # ─── Sheet 6: Cut Candidates ───────────────────────
    ws6 = wb.create_sheet("Cut Candidates")

    cut_headers = ["Chapter", "Words", "Grade", "Cut Rec", "Core Theme",
                    "Merge With", "Cut Reasoning", "Overall Score"]
    style_header(ws6, cut_headers)

    cut_row = 2
    # Sort: CUTTABLE first, then MERGEABLE, then ESSENTIAL
    cut_order = {"CUTTABLE": 0, "MERGEABLE": 1, "ESSENTIAL": 2}
    sorted_reviews = sorted(reviews.items(),
                             key=lambda x: cut_order.get(x[1].get("cut_recommendation", "?"), 3))
    for title, review in sorted_reviews:
        cut_rec = review.get("cut_recommendation", "?")
        cut_fill = {"ESSENTIAL": essential_fill, "MERGEABLE": mergeable_fill, "CUTTABLE": cuttable_fill}.get(cut_rec, None)
        data = [
            title,
            review.get("word_count", 0),
            review.get("chapter_grade", "?"),
            cut_rec,
            review.get("core_theme", "")[:60],
            review.get("merge_candidate") or "—",
            review.get("cut_reasoning", "")[:150],
            review.get("overall_score", 0),
        ]
        for col, val in enumerate(data, 1):
            cell = ws6.cell(row=cut_row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 4 and cut_fill:
                cell.fill = cut_fill
        cut_row += 1

    # Summary row
    cuttable_count = sum(1 for _, r in reviews.items() if r.get("cut_recommendation") == "CUTTABLE")
    mergeable_count = sum(1 for _, r in reviews.items() if r.get("cut_recommendation") == "MERGEABLE")
    cuttable_words = sum(r.get("word_count", 0) for _, r in reviews.items() if r.get("cut_recommendation") == "CUTTABLE")
    cut_row += 1
    ws6.cell(row=cut_row, column=1, value="SUMMARY").font = Font(bold=True)
    ws6.cell(row=cut_row, column=4, value=f"{cuttable_count} cuttable, {mergeable_count} mergeable")
    ws6.cell(row=cut_row, column=2, value=f"{cuttable_words} words cuttable")

    ws6.column_dimensions["A"].width = 35
    ws6.column_dimensions["B"].width = 8
    ws6.column_dimensions["C"].width = 7
    ws6.column_dimensions["D"].width = 11
    ws6.column_dimensions["E"].width = 30
    ws6.column_dimensions["F"].width = 30
    ws6.column_dimensions["G"].width = 50
    ws6.column_dimensions["H"].width = 12

    # ─── Sheet 7: Repetition Map ───────────────────────
    ws7 = wb.create_sheet("Repetition Map")

    if repetition_analysis and repetition_analysis.get("theme_clusters"):
        rep_headers = ["Theme Cluster", "Chapters", "Severity", "Recommendation"]
        style_header(ws7, rep_headers)

        rep_row = 2
        for cluster in repetition_analysis.get("theme_clusters", []):
            sev = cluster.get("severity", "?")
            sev_fill = {"HIGH": PatternFill("solid", fgColor="F4CCCC"),
                        "MEDIUM": PatternFill("solid", fgColor="FFF2CC"),
                        "LOW": PatternFill("solid", fgColor="E2EFDA")}.get(sev, None)
            data = [
                cluster.get("theme", ""),
                ", ".join(cluster.get("chapters", [])),
                sev,
                cluster.get("recommendation", ""),
            ]
            for col, val in enumerate(data, 1):
                cell = ws7.cell(row=rep_row, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if col == 3 and sev_fill:
                    cell.fill = sev_fill
            rep_row += 1

        # Summary
        rep_row += 1
        ws7.cell(row=rep_row, column=1, value="SUMMARY").font = Font(bold=True)
        ws7.cell(row=rep_row, column=2, value=repetition_analysis.get("summary", ""))
        rep_row += 1
        ws7.cell(row=rep_row, column=1, value="Unique chapters")
        ws7.cell(row=rep_row, column=2, value=", ".join(repetition_analysis.get("unique_chapters", [])))
        rep_row += 1
        ws7.cell(row=rep_row, column=1, value="Reducible chapters")
        ws7.cell(row=rep_row, column=2, value=repetition_analysis.get("total_reducible", 0))
        rep_row += 1
        ws7.cell(row=rep_row, column=1, value="Recommended final count")
        ws7.cell(row=rep_row, column=2, value=repetition_analysis.get("recommended_final_count", "?"))

        ws7.column_dimensions["A"].width = 35
        ws7.column_dimensions["B"].width = 50
        ws7.column_dimensions["C"].width = 10
        ws7.column_dimensions["D"].width = 60
    else:
        ws7.cell(row=1, column=1, value="Cross-chapter repetition analysis not available")

    # Save
    wb.save(output_path)
    print(f"\n  Saved: {output_path}")
    return output_path


# ═══════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Final Editor — Non-destructive review")
    parser.add_argument("--manuscript", required=True, help="Path to .docx manuscript")
    parser.add_argument("--output", default="", help="Output xlsx path (default: Final_Review.xlsx)")
    parser.add_argument("--chapters", default="", help="Comma-separated chapter numbers to review (default: all)")
    parser.add_argument("--anthropic-key", default=os.environ.get("ANTHROPIC_API_KEY", ""))
    parser.add_argument("--model", default="claude-sonnet-4-20250514", help="Review model (default: Sonnet)")
    args = parser.parse_args()

    api_key = args.anthropic_key
    if not api_key:
        print("ERROR: Set ANTHROPIC_API_KEY")
        sys.exit(1)

    output_path = args.output or str(SCRIPT_DIR / "Final_Review.xlsx")

    # Extract chapters
    print("=" * 64)
    print("  FINAL EDITOR — NON-DESTRUCTIVE REVIEW")
    print(f"  Manuscript: {args.manuscript}")
    print(f"  Model: {args.model}")
    print("=" * 64)
    print()
    print("  Extracting chapters...")

    chapters = extract_chapters_from_docx(args.manuscript)
    if not chapters:
        print("  ERROR: No chapters found")
        sys.exit(1)

    print(f"  Found {len(chapters)} chapters")

    # Filter if specified
    if args.chapters:
        filter_nums = [n.strip() for n in args.chapters.split(",")]
        filtered = {}
        for title, text in chapters.items():
            for num in filter_nums:
                if f"Chapter {num}:" in title or f"chapter {num}:" in title.lower():
                    filtered[title] = text
        if filtered:
            chapters = filtered
            print(f"  Filtered to {len(chapters)} chapters: {list(chapters.keys())}")
        else:
            print(f"  WARNING: No chapters matched filter {filter_nums}, reviewing all")

    # Review each chapter
    reviews = {}
    total = len(chapters)

    for idx, (title, text) in enumerate(chapters.items(), 1):
        words = len(text.split())
        print(f"\n  [{idx}/{total}] {title} ({words} words)")
        print(f"    Reviewing...")

        t0 = time.time()
        review = review_chapter(title, text, api_key, args.model)
        elapsed = time.time() - t0

        grade = review.get("chapter_grade", "?")
        overall = review.get("overall_score", 0)
        issues = len(review.get("issues", []))
        high_issues = sum(1 for i in review.get("issues", []) if i.get("severity") == "HIGH")
        meme_count = review.get("meme_quotes_found", 0)

        print(f"    Grade: {grade} | Overall: {overall}/10 | Issues: {issues} ({high_issues} HIGH) | Meme quotes: {meme_count} | {elapsed:.0f}s")

        cut_rec = review.get("cut_recommendation", "?")
        core_theme = review.get("core_theme", "")[:50]
        field_score = review.get("field_story_score", 0)
        field_stories = review.get("field_stories", [])
        weak_stories = sum(1 for s in field_stories if s.get("landing") in ("WEAK", "SUMMARY"))

        if review.get("best_line"):
            print(f"    Best: \"{review['best_line'][:80]}\"")
        print(f"    Cut: {cut_rec} | Theme: {core_theme} | Field stories: {field_score}/10 ({weak_stories} weak)")
        if cut_rec == "CUTTABLE":
            print(f"    >>> CUT CANDIDATE: {review.get('cut_reasoning', '')[:80]}")
        if cut_rec == "MERGEABLE":
            print(f"    >>> MERGE WITH: {review.get('merge_candidate', '?')} — {review.get('cut_reasoning', '')[:60]}")
        if high_issues > 0:
            for issue in review.get("issues", []):
                if issue.get("severity") == "HIGH":
                    print(f"    HIGH: {issue.get('issue', '')[:80]}")

        reviews[title] = review
        time.sleep(0.5)  # Rate limit buffer

    # Cross-chapter repetition analysis
    print(f"\n{'=' * 64}")
    print(f"  Running cross-chapter repetition analysis...")
    repetition = analyze_cross_chapter_repetition(reviews, api_key, args.model)

    if repetition.get("theme_clusters"):
        print(f"  Found {len(repetition['theme_clusters'])} theme clusters:")
        for cluster in repetition["theme_clusters"]:
            sev = cluster.get("severity", "?")
            theme = cluster.get("theme", "?")
            chapters = ", ".join(cluster.get("chapters", []))
            print(f"    [{sev}] \"{theme}\" — {chapters}")
        print(f"  Reducible: {repetition.get('total_reducible', 0)} chapters")
        print(f"  Recommended final count: {repetition.get('recommended_final_count', '?')}")
    else:
        print(f"  No significant theme clusters found")

    # Build spreadsheet
    print(f"\n{'=' * 64}")
    print(f"  Building review spreadsheet...")
    build_review_spreadsheet(reviews, output_path, repetition_analysis=repetition)

    # Summary
    all_issues = sum(len(r.get("issues", [])) for r in reviews.values())
    high_total = sum(sum(1 for i in r.get("issues", []) if i.get("severity") == "HIGH") for r in reviews.values())
    avg_score = sum(r.get("overall_score", 0) for r in reviews.values()) / len(reviews) if reviews else 0
    cuttable = sum(1 for r in reviews.values() if r.get("cut_recommendation") == "CUTTABLE")
    mergeable = sum(1 for r in reviews.values() if r.get("cut_recommendation") == "MERGEABLE")
    cuttable_words = sum(r.get("word_count", 0) for r in reviews.values() if r.get("cut_recommendation") == "CUTTABLE")
    weak_field = sum(sum(1 for s in r.get("field_stories", []) if s.get("landing") in ("WEAK", "SUMMARY")) for r in reviews.values())

    print(f"\n  REVIEW COMPLETE")
    print(f"  Chapters reviewed: {len(reviews)}")
    print(f"  Average score: {avg_score:.1f}/10")
    print(f"  Total issues: {all_issues} ({high_total} HIGH)")
    print(f"  Cut candidates: {cuttable} ({cuttable_words} words)")
    print(f"  Merge candidates: {mergeable}")
    print(f"  Weak field stories: {weak_field}")
    print(f"  Output: {output_path}")
    print(f"{'=' * 64}")

    # Also save raw JSON
    json_path = output_path.replace(".xlsx", ".json")
    with open(json_path, "w") as f:
        json.dump(reviews, f, indent=2, default=str)
    print(f"  Raw JSON: {json_path}")


if __name__ == "__main__":
    main()
